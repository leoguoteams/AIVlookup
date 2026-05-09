import json
import os
import threading
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import numpy as np
import openpyxl
from api_manager import APIManager, KeyObfuscator
from matcher_logic import MatcherLogic

CONFIG_FILE = "config.json"
VECTOR_CACHE_DIR = "vector_cache"
os.makedirs(VECTOR_CACHE_DIR, exist_ok=True)

def save_log(content):
    with open("run_log.log", "a", encoding="utf-8") as f:
        f.write(f"[{time.strftime('%H:%M:%S')}] {content}\n")

def log_print(widget, content):
    widget.config(state=tk.NORMAL)
    widget.insert(tk.END, f"[{time.strftime('%H:%M:%S')}] {content}\n")
    widget.see(tk.END)
    widget.config(state=tk.DISABLED)
    save_log(content)

class ColumnMapper(tk.Toplevel):
    """Dialog to map Excel columns to required fields."""
    def __init__(self, parent, file_path, required_fields, title="Map Columns"):
        super().__init__(parent)
        self.title(title)
        self.geometry("400x300")
        self.result = None
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            self.columns = [str(cell) for cell in ws[1]]
        except Exception as e:
            messagebox.showerror("Error", f"Failed to read file: {e}")
            self.destroy()
            return

        self.mappings = {}
        for i, field in enumerate(required_fields):
            frame = ttk.Frame(self)
            frame.pack(fill='x', padx=20, pady=5)
            ttk.Label(frame, text=f"{field}:", width=15).pack(side='left')
            combo = ttk.Combobox(frame, values=self.columns, state="readonly")
            combo.pack(side='left', fill='x', expand=True)
            # Auto-suggest based on keyword
            for col in self.columns:
                if field.lower() in col.lower():
                    combo.set(col)
                    break
            self.mappings[field] = combo

        btn_frame = ttk.Frame(self)
        btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="Confirm", command=self.confirm).pack()

    def confirm(self):
        self.result = {f: c.get() for f, c in self.mappings.items()}
        self.destroy()

class ResultAuditWindow(tk.Toplevel):
    """Window for auditing and refining match results."""
    def __init__(self, parent, results):
        super().__init__(parent)
        self.title("Result Audit & Review")
        self.geometry("1000x600")
        self.results = results # List of [A, B_Name, B_Spec, B_Code, Score, Source, Grade, Status]
        
        # Filter variables
        self.filter_grade = tk.StringVar(value="All")
        
        self.create_widgets()

    def create_widgets(self):
        top_frame = ttk.Frame(self)
        top_frame.pack(fill='x', padx=10, pady=10)
        
        ttk.Label(top_frame, text="Filter by Grade:").pack(side='left')
        combo = ttk.Combobox(top_frame, textvariable=self.filter_grade, values=["All", "High", "Medium", "Low"], state="readonly", width=10)
        combo.pack(side='left', padx=5)
        combo.bind("<<ComboboxSelected>>", lambda e: self.refresh_table())
        
        # Table
        self.tree = ttk.Treeview(self, columns=("A", "B", "Score", "Grade", "Status"), show='headings')
        self.tree.heading("A", text="Target (A)")
        self.tree.heading("B", text="Matched (B)")
        self.tree.heading("Score", text="Score")
        self.tree.heading("Grade", text="Grade")
        self.tree.heading("Status", text="Status")
        self.tree.column("Score", width=70)
        self.tree.column("Grade", width=70)
        self.tree.column("Status", width=100)
        self.tree.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.refresh_table()

    def refresh_table(self):
        for i in self.tree.get_children():
            self.tree.delete(i)
            
        filter_val = self.filter_grade.get()
        for row in self.results:
            if filter_val == "All" or row[6] == filter_val:
                self.tree.insert("", tk.END, values=(row[0], row[1], row[4], row[6], row[7]))

class AISemanticMatchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("AI 语义智能匹配工具 · 专业版 v1.2")
        self.root.geometry("900x720")
        self.root.resizable(False, False)

        self.api_key = tk.StringVar()
        self.a_path = tk.StringVar()
        self.b_path = tk.StringVar()
        self.score_threshold = tk.IntVar(value=30)
        
        # Configuration storage
        self.b_config = {} # Store mapped columns for B table
        self.a_config = {} # Store mapped column for A table
        
        self.b_embeddings = None
        self.b_data = None
        self.stop_flag = threading.Event()
        
        self.load_config()
        self.create_widgets()

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                    # Deobfuscate API key
                    self.api_key.set(KeyObfuscator.deobfuscate(cfg.get("api_key", "")))
            except:
                pass

    def save_config(self):
        try:
            # Obfuscate API key
            obfuscated_key = KeyObfuscator.obfuscate(self.api_key.get())
            cfg = {"api_key": obfuscated_key}
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(cfg, f, indent=2)
            messagebox.showinfo("成功", "配置已保存")
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {e}")

    def validate_api(self):
        key = self.api_key.get().strip()
        if not key:
            messagebox.showwarning("提示", "请输入 API Key")
            return
        
        log_print(self.log_box, "正在校验 API...")
        api_mgr = APIManager(key)
        if api_mgr.get_embeddings(["test"]) is not None:
            log_print(self.log_box, "✅ API 校验通过")
            self.api_status_label.config(foreground="green")
        else:
            log_print(self.log_box, "❌ API 校验失败")
            self.api_status_label.config(foreground="red")

    def select_a(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.a_path.set(p)
            mapper = ColumnMapper(self.root, p, ["Target Text"], "Map A Table Column")
            self.root.wait_window(mapper)
            if mapper.result:
                self.a_config = mapper.result
                log_print(self.log_box, f"✅ A表映射: {self.a_config}")

    def select_b(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.b_path.set(p)
            mapper = ColumnMapper(self.root, p, ["Code", "Name", "Spec", "Size"], "Map B Table Columns")
            self.root.wait_window(mapper)
            if mapper.result:
                self.b_config = mapper.result
                log_print(self.log_box, f"✅ B表映射: {self.b_config}")
                self.b_embeddings = None
                self.vector_status_label.config(foreground="red")

    def start_b_vectorization(self):
        if not self.b_path.get() or not self.b_config:
            messagebox.showwarning("提示", "请选择B表并完成列映射")
            return
        self.stop_flag.clear()
        self.vectorizing = True
        self.vectorize_btn.config(text="停止向量化")
        threading.Thread(target=self.prepare_b_embeddings, daemon=True).start()

    def prepare_b_embeddings(self):
        path = self.b_path.get()
        cache_path = os.path.join(VECTOR_CACHE_DIR, f"{os.path.basename(path)}.npy")
        
        try:
            log_print(self.log_box, "正在读取 B 表并生成向量...")
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            
            # Use mapped columns
            data = []
            cols = {k: v for k, v in self.b_config.items()} # Field name -> Col Name
            
            # Find indices of the mapped column names
            header = [str(c) for c in ws[1]]
            idx_map = {field: header.index(col) for field, col in cols.items()}
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[idx_map["Code"]]) if len(row) > idx_map["Code"] else ""
                name = str(row[idx_map["Name"]]) if len(row) > idx_map["Name"] else ""
                spec = str(row[idx_map["Spec"]]) if len(row) > idx_map["Spec"] else ""
                size = str(row[idx_map["Size"]]) if len(row) > idx_map["Size"] else ""
                txt = (name + spec + size).strip()
                if txt:
                    data.append((code, name, spec, size, txt))

            # Vectorize using APIManager
            api_mgr = APIManager(self.api_key.get().strip())
            texts = [d[4] for d in data]
            self.b_embeddings = api_mgr.get_embeddings(texts)
            self.b_data = data
            
            if self.b_embeddings is not None:
                np.save(cache_path, {"vectors": self.b_embeddings, "data": data})
                log_print(self.log_box, f"✅ 向量化成功: {len(data)} 条")
                self.vector_status_label.config(foreground="blue")
            else:
                log_print(self.log_box, "❌ 向量生成失败")
                
        except Exception as e:
            log_print(self.log_box, f"❌ 错误: {e}")
        finally:
            self.vectorizing = False
            self.vectorize_btn.config(text="开始向量化")

    def stop_vectorization(self):
        self.stop_flag.set()
        log_print(self.log_box, "停止向量化请求...")

    def start_match(self):
        if not self.a_path.get() or not self.a_config:
            messagebox.showwarning("提示", "请选择A表并完成列映射")
            return
        if self.b_embeddings is None:
            messagebox.showwarning("提示", "请先完成B表向量化")
            return
        threading.Thread(target=self.run_match, daemon=True).start()

    def run_match(self):
        log_print(self.log_box, "🚀 开始专业语义匹配 (v1.2)...")
        api_mgr = APIManager(self.api_key.get().strip())
        matcher = MatcherLogic(api_mgr, self.score_threshold.get())
        
        try:
            wb = openpyxl.load_workbook(self.a_path.get(), data_only=True)
            ws = wb.active
            header = [str(c) for c in ws[1]]
            a_col_name = self.a_config["Target Text"]
            a_idx = header.index(a_col_name)
            
            a_texts = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                a_texts.append(str(row[a_idx]) if len(row) > a_idx else "")

            results = []
            for i, text in enumerate(a_texts):
                if self.stop_flag.is_set(): break
                log_print(self.log_box, f"Matching {i+1}/{len(a_texts)}: {text[:20]}...")
                res = matcher.match_single_item(text, self.b_data, self.b_embeddings)
                # format: (B_code, B_name, B_spec, B_size, Score, Source, Grade, Status)
                results.append([text, res[1], res[2], res[0], res[4], res[5], res[6], res[7]])
            
            # Save to Excel
            out_path = os.path.join(os.path.dirname(self.a_path.get()), "AI匹配结果_v1.2.xlsx")
            wb_out = openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.append(["A表内容", "B表名称", "B表规格", "B表编码", "AI得分", "来源", "信心等级", "状态"])
            for r in results:
                ws_out.append(r)
            wb_out.save(out_path)
            
            log_print(self.log_box, f"✅ 匹配完成！结果保存至: {out_path}")
            
            # Open Audit Window
            self.root.after(0, lambda: ResultAuditWindow(self.root, results))
            messagebox.showinfo("完成", "匹配完成，已打开审核窗口")
            
        except Exception as e:
            log_print(self.log_box, f"❌ 匹配过程中出错: {e}")

    def create_widgets(self):
        # API Config
        f1 = ttk.LabelFrame(self.root, text="API 配置")
        f1.place(x=20, y=10, width=860, height=80)
        ttk.Label(f1, text="API Key:").place(x=10, y=25)
        self.api_key_entry = tk.Entry(f1, textvariable=self.api_key, width=40, show="*")
        self.api_key_entry.place(x=100, y=23)
        ttk.Button(f1, text="校验 API", command=self.validate_api).place(x=520, y=22)
        ttk.Button(f1, text="保存配置", command=self.save_config).place(x=620, y=22)
        self.api_status_label = ttk.Label(f1, text="●", font=("Arial", 14))
        self.api_status_label.place(x=800, y=20)
        self.api_status_label.config(foreground="red")

        # B Table
        f2 = ttk.LabelFrame(self.root, text="第一步：B表处理 (底表向量化)")
        f2.place(x=20, y=100, width=860, height=160)
        ttk.Label(f2, text="B表文件:").place(x=10, y=20)
        ttk.Entry(f2, textvariable=self.b_path, width=50).place(x=80, y=18)
        ttk.Button(f2, text="选择并映射", command=self.select_b).place(x=570, y=17)
        
        self.vectorize_btn = ttk.Button(f2, text="开始向量化", command=self.start_b_vectorization)
        self.vectorize_btn.place(x=660, y=17)
        self.vector_status_label = ttk.Label(f2, text="●", font=("Arial", 14))
        self.vector_status_label.place(x=750, y=16)
        self.vector_status_label.config(foreground="red")
        
        ttk.Label(f2, text="提示: 请确保B表包含 [编码, 名称, 规格, 尺寸] 相关列").place(x=10, y=60)
        
        # A Table
        f3 = ttk.LabelFrame(self.root, text="第二步：A表处理 (执行语义匹配)")
        f3.place(x=20, y=270, width=860, height=160)
        ttk.Label(f3, text="A表文件:").place(x=10, y=20)
        ttk.Entry(f3, textvariable=self.a_path, width=50).place(x=80, y=18)
        ttk.Button(f3, text="选择并映射", command=self.select_a).place(x=570, y=17)
        
        ttk.Label(f3, text="匹配阈值:").place(x=10, y=60)
        ttk.Scale(f3, variable=self.score_threshold, from_=0, to=100, orient=tk.HORIZONTAL).place(x=80, y=58, width=200)
        ttk.Label(f3, textvariable=self.score_threshold).place(x=290, y=60)
        
        self.match_btn = ttk.Button(f3, text="开始匹配", command=self.start_match, width=20)
        self.match_btn.place(x=400, y=55)

        # Log
        self.log_box = tk.Text(self.root, state=tk.DISABLED, font=("Consolas", 10))
        self.log_box.place(x=20, y=440, width=860, height=250)

if __name__ == "__main__":
    root = tk.Tk()
    app = AISemanticMatchApp(root)
    root.mainloop()

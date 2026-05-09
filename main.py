import json
import os
import sys
import subprocess
import threading
import time
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import math
import requests
import urllib3
import numpy as np
import openpyxl
from concurrent.futures import ThreadPoolExecutor
from openpyxl.utils.exceptions import InvalidFileException

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ==================== 【阿里云官方正确接口地址】已修复 ====================
CONFIG_FILE = "config.json"
VECTOR_CACHE_DIR = "vector_cache"
os.makedirs(VECTOR_CACHE_DIR, exist_ok=True)

EMBED_URL = "https://dashscope.aliyuncs.com/api/v1/services/embeddings/text-embedding/text-embedding"
CHAT_URL = "https://dashscope.aliyuncs.com/api/v1/services/aigc/multimodal-generation/generation"

# ==================== 工具函数 ====================
def clean_text(text):
    if not text:
        return ""
    text = str(text).strip()
    text = text.replace("　", "").replace(" ", "")
    return text

def save_log(content):
    with open("run_log.log", "a", encoding="utf-8") as f:
        f.write(f"[{time.strftime('%H:%M:%S')}] {content}\n")

def log_print(widget, content):
    widget.config(state=tk.NORMAL)
    widget.insert(tk.END, f"[{time.strftime('%H:%M:%S')}] {content}\n")
    widget.see(tk.END)
    widget.config(state=tk.DISABLED)
    save_log(content)

# ==================== API 函数（带重试） ====================
def get_embeddings_with_retry(text_list, api_key, max_retry=3):
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    data = {
        "model": "text-embedding-v3",
        "input": {"texts": text_list}
    }
    for attempt in range(max_retry):
        try:
            res = requests.post(EMBED_URL, json=data, headers=headers, verify=False, timeout=30)
            if res.status_code == 200:
                return np.array([item["embedding"] for item in res.json()["output"]["embeddings"]])
            else:
                print(f"API请求失败，状态码: {res.status_code}, 响应: {res.text}")
                time.sleep(1)
        except Exception as e:
            print(f"API请求异常: {str(e)}")
            time.sleep(1)
    return None

def get_embeddings_parallel(texts, api_key, batch_size=40, max_concurrency=4, progress_callback=None, stop_check=None):
    """并行获取嵌入向量，使用信号量限制并发数"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    import threading

    semaphore = threading.Semaphore(max_concurrency)
    results = [None] * len(texts)  # 预分配结果列表，保持顺序

    def fetch_batch(batch_idx, batch_texts):
        if stop_check and stop_check():
            return
        with semaphore:
            emb = get_embeddings_with_retry(batch_texts, api_key)
            if emb is not None:
                for j, e in enumerate(emb):
                    results[batch_idx * batch_size + j] = e
            else:
                print(f"批次 {batch_idx} 向量生成失败")
            if progress_callback:
                progress_callback(len(batch_texts))

    # 分批
    batches = []
    for i in range(0, len(texts), batch_size):
        if stop_check and stop_check():
            break
        batch_texts = texts[i:i+batch_size]
        batch_idx = i // batch_size
        batches.append((batch_idx, batch_texts))

    with ThreadPoolExecutor(max_workers=max_concurrency) as executor:
        futures = [executor.submit(fetch_batch, idx, txts) for idx, txts in batches]
        for f in as_completed(futures):
            pass  # 等待完成

    # 检查是否有失败
    failed_count = sum(1 for r in results if r is None)
    if failed_count > 0:
        print(f"⚠️ 警告: {failed_count}/{len(texts)} 条嵌入失败")
    results_array = np.array(results)
    # 防御：如果返回 1D 说明生成失败，返回 None 让调用方处理
    if results_array.ndim != 2 or results_array.shape[0] == 0:
        return None
    return results_array

def llm_find_best_from_top3(a_text, top3_data, api_key, log_func=print):
    """让LLM从Top3候选中选出最相似的1条"""
    b_list_str = ""
    for i, item in enumerate(top3_data):
        b_list_str += f"[{i}] 名称:{item[1]}, 规格:{item[2]}, 尺寸:{item[3]}\n"
    # Intentionally use string concatenation (not .format/f-string templating) for this prompt:
    # the template contains literal JSON braces, and dynamic text may also include brace characters.
    # Concatenation keeps all braces as plain text and avoids fragile formatter parsing issues.
    prompt = ("""##角色:你是物料匹配专家
##任务:从"##B候选"的3条物料中找出与"##A材料"最相似的1条，并给出相似度评分
##A材料: """ + a_text + """
##B候选:
""" + b_list_str + """
##匹配标准:
- 只有名称/规格/尺寸中有一项语义非常近才匹配
- 必须给出一个相似度评分(0-100整数)
  示例：
  {
  "top3": [
    {"index": 0, "score": 85},
    {"index": 1, "score": 70},
    {"index": 2, "score": 60}
  ],
  "best_index": 0
}
##输出格式:
- 返回Top3分数最高的一个JSON，示例格式如下：{"index": 0, "score": 85}
"""
)
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    data = {
        "model": "qwen3.5-plus",
        "input": {
            "messages": [{"role": "user", "content": prompt}]
        },
        "parameters": {"temperature": 0.5, "result_format": "message"}
    }
    try:
        #log_func(f"[LLM请求] URL: {CHAT_URL}")
        #log_func(f"[LLM请求] Model: qwen3.5-plus")
        #log_func(f"[LLM请求] Prompt:\n{prompt[:500]}..." 
        #if len(prompt) > 500 
        #else f"[LLM请求] Prompt:\n{prompt}")
        res = requests.post(CHAT_URL, json=data, headers=headers, verify=False, timeout=60)
        #log_func(f"[LLM响应] 状态码: {res.status_code}")
        if res.status_code == 200:
            resp_json = res.json()
            log_func(f"[LLM响应] 原始响应: {resp_json}")
            txt = None
            if "output" in resp_json and "choices" in resp_json["output"]:
                content = resp_json["output"]["choices"][0]["message"]["content"]
                # content可能是 [{"text": "..."}] 或直接的字符串
                if isinstance(content, list) and len(content) > 0:
                    first_item = content[0]
                    if isinstance(first_item, dict):
                        txt = first_item.get('text', '')
                    else:
                        txt = str(first_item)
                elif isinstance(content, str):
                    txt = content
            elif "output" in resp_json and "text" in resp_json["output"]:
                txt = resp_json["output"]["text"]
            log_func(f"[LLM响应] 提取文本: '{txt}'")
            if txt:
                import re
                import json
                txt_str = str(txt)
                # First, try to parse the whole content as JSON to avoid regex fragility
                try:
                    parsed = json.loads(txt_str)
                    if isinstance(parsed, dict):
                        if "top3" in parsed:
                            top3_list = parsed["top3"]
                            if isinstance(top3_list, list) and len(top3_list) > 0:
                                best = max(top3_list, key=lambda x: int(x.get("score", 0)))
                                best_idx = int(best.get("index", -1))
                                if 0 <= best_idx < len(top3_data):
                                    item = top3_data[best_idx]
                                    return (best_idx, item[0], item[1], item[2], item[3], int(best.get("score", 0)))
                        # Old style: {"index":0, "score":85}
                        idx = int(parsed.get('index', -1))
                        score = int(parsed.get('score', 0))
                        if idx >= 0 and idx < len(top3_data):
                            item = top3_data[idx]
                            return (idx, item[0], item[1], item[2], item[3], score)
                except Exception:
                    pass
                # 尝试提取JSON格式 {"index": 0, "score": 85}
                json_match = re.search(r'\{[^}]+\}', txt_str)
                if json_match:
                    try:
                        data = json.loads(json_match.group())
                        # 新的多分数格式：{ "top3": [ {"index":0,"score":85}, ... ], "best_index":0 }
                        if isinstance(data, dict) and "top3" in data:
                            top3_list = data["top3"]
                            if isinstance(top3_list, list) and len(top3_list) > 0:
                                best = max(top3_list, key=lambda x: int(x.get("score", 0)))
                                best_idx = int(best.get("index", -1))
                                if 0 <= best_idx < len(top3_data):
                                    item = top3_data[best_idx]
                                    return (best_idx, item[0], item[1], item[2], item[3], int(best.get("score", 0)))
                                return (-1, None, None, None, None, 0)
                        # 旧格式：{ "index": 0, "score": 85 }
                        idx = data.get('index', -1)
                        score = data.get('score', 0)
                        log_func(f"[LLM响应] 索引: {idx}, 评分: {score}")
                        if idx >= 0 and idx < len(top3_data):
                            item = top3_data[idx]
                            return (idx, item[0], item[1], item[2], item[3], score)
                        # Fallback: attempt to use best_index/index if provided
                        try:
                            fallback_idx = int(data.get("best_index", data.get("index", -1)))
                            if 0 <= fallback_idx < len(top3_data):
                                item = top3_data[fallback_idx]
                                fallback_score = int(data.get("score", 0))
                                return (fallback_idx, item[0], item[1], item[2], item[3], fallback_score)
                        except Exception:
                            pass
                        return (-1, None, None, None, None, 0)
                    except:
                        pass
                # 备用：只提取数字索引
                num_match = re.search(r'[-]?\d+', txt_str)
                if num_match:
                    result = int(num_match.group())
                    log_func(f"[LLM响应] 返回索引: {result}")
                    if result >= 0 and result < len(top3_data):
                        item = top3_data[result]
                        return (result, item[0], item[1], item[2], item[3], 100)
                    return (-1, None, None, None, None, 0)
        else:
            log_func(f"[LLM响应] 请求失败: {res.text[:500]}")
    except Exception as e:
        log_func(f"[LLM异常] {e}")
    return (-1, None, None, None, None, 0)


# ==================== 主界面 ====================
class AISemanticMatchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("AI 语义智能匹配工具 · 企业专业版")
        self.root.geometry("900x720")
        self.root.resizable(False, False)

        # 变量
        self.api_key = tk.StringVar()
        self.a_path = tk.StringVar()
        self.b_path = tk.StringVar()
        self.score_threshold = tk.IntVar(value=30)
        self.a_col = tk.StringVar(value="B")
        self.use_cache = tk.BooleanVar(value=True)  # 是否使用已有缓存
        self.force_revectorize = tk.BooleanVar(value=False)  # 是否强制重新向量化
        # 互斥逻辑：使用缓存和强制重向量只能选一个
        self.use_cache.trace_add('write', lambda *args: (self.force_revectorize.set(False), self.update_vectorize_btn_state(), self.update_match_btn_state()) if self.use_cache.get() else (self.update_vectorize_btn_state(), self.update_match_btn_state()))
        self.force_revectorize.trace_add('write', lambda *args: (self.use_cache.set(False), self.update_vectorize_btn_state(), self.update_match_btn_state()) if self.force_revectorize.get() else (self.update_vectorize_btn_state(), self.update_match_btn_state()))
        self.batch_size = tk.IntVar(value=10)         # 批量大小，默认10
        self.max_concurrency = tk.IntVar(value=1)      # 最大并发数，默认1
        # 新增：LLM 并发控制，策略 A（并发发送LLM请求）
        self.llm_concurrency = tk.IntVar(value=4)      # 默认4

        self.b_data = None
        self.b_embeddings = None
        self.cache_file = None
        # Threshold gates (configurable defaults)
        self.sim_top1_th = 0.90
        self.diff_top12_th = 0.035
        # UI: threshold advanced controls (will be wired to config.json)
        self.use_default_thresholds = tk.BooleanVar(value=False)
        self.load_config()

        # UI
        self.stop_flag = threading.Event()  # 用于控制停止运行
        self.vectorization_complete = tk.StringVar(value="red")  # 默认红灯，表示未完成
        self.vectorizing = tk.BooleanVar(value=False)  # 是否正在向量化
        self.create_widgets()
        # After UI is created, synchronize thresholds with config (if needed)
        self.sync_thresholds_with_config()
        # 监听文件路径变化，更新匹配按钮状态
        self.b_path.trace_add('write', lambda *args: self.update_match_btn_state())
        self.a_path.trace_add('write', lambda *args: self.update_match_btn_state())

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                cfg = json.load(open(CONFIG_FILE, "r", encoding="utf-8"))
                self.api_key.set(cfg.get("api_key", ""))
                self.batch_size.set(cfg.get("batch_size", 10))
                self.max_concurrency.set(cfg.get("max_concurrency", 1))
                self.use_cache.set(cfg.get("use_cache", True))
                # LL.M concurrency setting from config (single consolidated path)
                if hasattr(self, 'llm_concurrency'):
                    self.llm_concurrency.set(cfg.get("llm_concurrency", self.llm_concurrency.get()))
                # Threshold gates (from config, fallback to defaults)
                self.sim_top1_th = cfg.get("sim_top1_th", self.sim_top1_th if hasattr(self, 'sim_top1_th') else 0.90)
                self.diff_top12_th = cfg.get("diff_top12_th", self.diff_top12_th if hasattr(self, 'diff_top12_th') else 0.035)
                # Threshold defaults override by config
                if 'use_default_thresholds' in cfg:
                    self.use_default_thresholds.set(cfg.get("use_default_thresholds", False))
                if cfg.get("use_default_thresholds", False):
                    self.sim_top1_th = 0.90
                    self.diff_top12_th = 0.035
                    self.sim_top1_slider.set(self.sim_top1_th)
                    self.diff_top12_slider.set(self.diff_top12_th)
                    self.sim_top1_label.config(text=f"{self.sim_top1_th:.3f}")
                    self.diff_top12_label.config(text=f"{self.diff_top12_th:.3f}")
                # 新增：权重配置（默认 0.5/0.5）
                self.llm_weight = cfg.get("llm_weight", 0.5)
                self.cosine_weight = cfg.get("cosine_weight", 0.5)
            except:
                pass

    def save_config(self):
        cfg = {
            "api_key": self.api_key.get(),
            "batch_size": self.batch_size.get(),
            "max_concurrency": self.max_concurrency.get(),
            "use_cache": self.use_cache.get(),
            "llm_concurrency": self.llm_concurrency.get(),
            "sim_top1_th": self.sim_top1_th,
            "diff_top12_th": self.diff_top12_th,
            "use_default_thresholds": self.use_default_thresholds.get(),
            "llm_weight": getattr(self, 'llm_weight', 0.5),
            "cosine_weight": getattr(self, 'cosine_weight', 0.5)
        }
        json.dump(cfg, open(CONFIG_FILE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
        messagebox.showinfo("成功", "配置已保存")

    # Thresholds UI handlers
    def on_sim_top1_th_change(self, val):
        try:
            v = float(val)
        except ValueError:
            return
        self.sim_top1_th = max(0.0, min(1.0, v))
        if hasattr(self, 'sim_top1_label'):
            self.sim_top1_label.config(text=f"{self.sim_top1_th:.3f}")

    def on_diff_top12_th_change(self, val):
        try:
            v = float(val)
        except ValueError:
            return
        self.diff_top12_th = max(0.0, min(0.5, v))
        if hasattr(self, 'diff_top12_label'):
            self.diff_top12_label.config(text=f"{self.diff_top12_th:.3f}")
        self.save_config()

    def toggle_use_default_thresholds(self):
        if self.use_default_thresholds.get():
            # Apply defaults
            self.sim_top1_th = 0.90
            self.diff_top12_th = 0.035
            if hasattr(self, 'sim_top1_slider'):
                self.sim_top1_slider.set(self.sim_top1_th)
            if hasattr(self, 'diff_top12_slider'):
                self.diff_top12_slider.set(self.diff_top12_th)
            if hasattr(self, 'sim_top1_label'):
                self.sim_top1_label.config(text=f"{self.sim_top1_th:.3f}")
            if hasattr(self, 'diff_top12_label'):
                self.diff_top12_label.config(text=f"{self.diff_top12_th:.3f}")
            self.save_config()
        else:
            # Enable sliders for custom thresholds; do not override values
            if hasattr(self, 'sim_top1_slider'):
                self.sim_top1_slider.config(state='normal')
            if hasattr(self, 'diff_top12_slider'):
                self.diff_top12_slider.config(state='normal')

    def save_thresholds(self):
        # Persist current threshold values to config.json
        cfg = {
            "sim_top1_th": self.sim_top1_th,
            "diff_top12_th": self.diff_top12_th,
            "llm_concurrency": self.llm_concurrency.get() if hasattr(self, 'llm_concurrency') else 4,
            "use_default_thresholds": self.use_default_thresholds.get()
        }
        try:
            json.dump(cfg, open(CONFIG_FILE, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
            log_print(self.log_box, "阈值已保存")
        except Exception as e:
            log_print(self.log_box, f"阈值保存失败: {e}")

    def sync_thresholds_with_config(self):
        """Synchronize UI thresholds with config.json on startup.
        If use_default_thresholds is true, apply defaults. Otherwise align UI sliders
        with the persisted sim_top1_th and diff_top12_th.
        """
        try:
            # Prefill from config if exists
            cfg = {}
            if os.path.exists(CONFIG_FILE):
                import json as _json
                cfg = _json.load(open(CONFIG_FILE, 'r', encoding='utf-8'))
            # Apply use_default_thresholds first
            use_def = bool(cfg.get("use_default_thresholds", False))
            if use_def:
                self.use_default_thresholds.set(True)
                self.sim_top1_th = 0.90
                self.diff_top12_th = 0.035
                if hasattr(self, 'sim_top1_slider'):
                    self.sim_top1_slider.set(self.sim_top1_th)
                if hasattr(self, 'diff_top12_slider'):
                    self.diff_top12_slider.set(self.diff_top12_th)
                if hasattr(self, 'sim_top1_label'):
                    self.sim_top1_label.config(text=f"{self.sim_top1_th:.3f}")
                if hasattr(self, 'diff_top12_label'):
                    self.diff_top12_label.config(text=f"{self.diff_top12_th:.3f}")
            else:
                # Align with persisted thresholds if present
                self.use_default_thresholds.set(cfg.get("use_default_thresholds", False))
                self.sim_top1_th = cfg.get("sim_top1_th", getattr(self, 'sim_top1_th', 0.90))
                self.diff_top12_th = cfg.get("diff_top12_th", getattr(self, 'diff_top12_th', 0.035))
                if hasattr(self, 'sim_top1_slider'):
                    self.sim_top1_slider.set(self.sim_top1_th)
                if hasattr(self, 'diff_top12_slider'):
                    self.diff_top12_slider.set(self.diff_top12_th)
                if hasattr(self, 'sim_top1_label'):
                    self.sim_top1_label.config(text=f"{self.sim_top1_th:.3f}")
                if hasattr(self, 'diff_top12_label'):
                    self.diff_top12_label.config(text=f"{self.diff_top12_th:.3f}")
        except Exception:
            pass

    def validate_api(self):
        key = self.api_key.get().strip()
        if not key:
            messagebox.showwarning("提示", "请输入API Key")
            return
        log_print(self.log_box, "正在校验API...")
        try:
            test = get_embeddings_with_retry(["check"], key)
            if test is not None:
                log_print(self.log_box, "✅ API 校验通过")
                self.api_status.set("green")  # 设置绿灯
            else:
                log_print(self.log_box, "❌ API 校验失败")
                self.api_status.set("red")  # 设置红灯
        except Exception as e:
            log_print(self.log_box, f"❌ 校验异常：{str(e)}")
            self.api_status.set("red")  # 设置红灯
        self.api_status_label.config(foreground=self.api_status.get())

    def toggle_api_key_visibility(self):
        if not hasattr(self, 'api_key_entry'):
            return
        if self.api_key_show.get():
            self.api_key_entry.config(show="")
        else:
            self.api_key_entry.config(show="*")

    def on_use_cache_toggle(self):
        """勾选/取消"使用已有缓存"时的处理"""
        if self.use_cache.get():
            # 加载缓存文件列表
            self._refresh_cache_list()
        else:
            # 清除缓存选择
            self.cache_combo.set('')
            self.b_embeddings = None
            self.b_data = None
        self.update_match_btn_state()

    def _refresh_cache_list(self):
        """刷新缓存文件列表"""
        if not os.path.exists(VECTOR_CACHE_DIR):
            self.cache_combo['values'] = []
            return
        files = [f for f in os.listdir(VECTOR_CACHE_DIR) if f.endswith('.npy')]
        self.cache_combo['values'] = files
        if files:
            self.cache_combo.current(0)
            self._load_cache_file(files[0])

    def on_cache_selected(self, event=None):
        """选择缓存文件时的处理"""
        selected = self.cache_combo.get()
        if selected:
            self._load_cache_file(selected)

    def _load_cache_file(self, filename):
        """加载指定的缓存文件（包含向量和B表数据）"""
        cache_path = os.path.join(VECTOR_CACHE_DIR, filename)
        try:
            cached = np.load(cache_path, allow_pickle=True)
            # 兼容旧格式（只保存向量）和新格式（包含向量和数据）
            if cached.ndim == 0:  # 新格式：字典
                cache_data = cached.item()
                self.b_embeddings = cache_data['vectors']
                self.b_data = cache_data['data']
                log_print(self.log_box, f"✅ 已加载缓存: {filename} ({self.b_embeddings.shape[0]} 条向量)")
            else:  # 旧格式：只保存了向量，需要用户重新提供B表或提示
                self.b_embeddings = cached
                self.b_data = None
                log_print(self.log_box, f"⚠️ 旧缓存格式，请重新向量化以支持完整功能")
                messagebox.showwarning("提示", "旧缓存格式，请在勾选'强制重新向量化'后重新生成缓存")
            self.vectorization_complete.set("blue")
            self.vector_status_label.config(foreground=self.vectorization_complete.get())
            self.update_match_btn_state()
        except Exception as e:
            log_print(self.log_box, f"❌ 缓存加载失败: {str(e)}")
            self.cache_combo.set('')

    def update_vectorize_btn_state(self):
        if not hasattr(self, 'vectorizing') or not hasattr(self, 'vectorize_btn'):
            return
        if self.vectorizing.get():
            self.vectorize_btn.config(state="normal", text="停止向量化")
        elif self.force_revectorize.get():
            self.vectorize_btn.config(state="normal", text="开始向量化")
        elif self.use_cache.get() and self.vectorization_complete.get() == "blue":
            self.vectorize_btn.config(state="normal", text="开始向量化")
        else:
            self.vectorize_btn.config(state="disabled", text="开始向量化")

    def update_match_btn_state(self):
        if not hasattr(self, 'match_btn'):
            return
        # 向量化完成（通过缓存或强制向量化）且A表已上传时启用
        if self.b_embeddings is not None and self.a_path.get():
            self.match_btn.config(state="normal")
        else:
            self.match_btn.config(state="disabled")

    def select_a(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.a_path.set(p)

    def select_b(self):
        p = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if p:
            self.b_path.set(p)
            self.b_embeddings = None
            self.b_data = None
            self.vectorization_complete.set("red")
            self.vector_status_label.config(foreground=self.vectorization_complete.get())
            log_print(self.log_box, "已选择B文件，请先执行B表向量化处理")

    def start_b_vectorization(self):
        if not self.b_path.get():
            messagebox.showwarning("提示", "请选择B表")
            return
        self.stop_flag.clear()  # 重置停止标志
        self.vectorizing.set(True)
        self.vectorize_btn.config(text="停止向量化")
        threading.Thread(target=self.prepare_b_embeddings, daemon=True).start()

    def toggle_vectorization(self):
        if self.vectorizing.get():
            # 正在运行，点击停止
            self.stop_flag.set()
            log_print(self.log_box, "❌ 正在停止向量化...")
        else:
            # 未运行，点击开始
            self.start_b_vectorization()

    def prepare_b_embeddings(self):
        if self.stop_flag.is_set():
            return
        path = self.b_path.get()
        if not path:
            log_print(self.log_box, "❌ 未选择B文件")
            return

        self.cache_file = self._get_cache_path(path)
        log_print(self.log_box, "正在加载存货档案...")

        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            data = []

            # 自动查找列索引
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            col_to_idx = {}
            for i, cell in enumerate(header_row):
                if cell:
                    cell_str = str(cell).strip()
                    if "存货编码" in cell_str:
                        col_to_idx["code"] = i
                    elif "存货名称" in cell_str:
                        col_to_idx["name"] = i
                    elif "规格型号" in cell_str:
                        col_to_idx["spec"] = i
                    elif "尺寸" in cell_str:
                        col_to_idx["size"] = i

            if not all(k in col_to_idx for k in ["code", "name", "spec", "size"]):
                missing = [k for k in ["code", "name", "spec", "size"] if k not in col_to_idx]
                log_print(self.log_box, f"❌ B表缺少必要列: {missing}")
                return

            log_print(self.log_box, f"✅ 自动识别列: 编码={col_to_idx['code']}, 名称={col_to_idx['name']}, 规格={col_to_idx['spec']}, 尺寸={col_to_idx['size']}")

            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[col_to_idx["code"]]) if len(row) > col_to_idx["code"] else ""
                name = str(row[col_to_idx["name"]]) if len(row) > col_to_idx["name"] else ""
                spec = str(row[col_to_idx["spec"]]) if len(row) > col_to_idx["spec"] else ""
                size = str(row[col_to_idx["size"]]) if len(row) > col_to_idx["size"] else ""
                txt = clean_text(name + spec + size)
                if txt:
                    data.append((code, name, spec, size, txt))
            self.b_data = data

            if len(data) == 0:
                log_print(self.log_box, "❌ B表未发现有效文本，无法向量化")
                return

            if os.path.exists(self.cache_file) and (self.force_revectorize.get() or not self.use_cache.get()):
                try:
                    os.remove(self.cache_file)
                    log_print(self.log_box, "✅ 已删除旧缓存，准备重新生成向量")
                except Exception as e:
                    log_print(self.log_box, f"⚠️ 无法删除旧缓存：{str(e)}")

            if os.path.exists(self.cache_file):
                try:
                    cached = np.load(self.cache_file, allow_pickle=True)
                    # 兼容旧格式
                    if cached.ndim == 0:  # 新格式：字典
                        cache_data = cached.item()
                        self.b_embeddings = cache_data['vectors']
                        self.b_data = cache_data['data']
                        log_print(self.log_box, f"✅ 向量缓存加载成功：{cache_data['vectors'].shape[0]} 条")
                    else:  # 旧格式：只保存了向量
                        self.b_embeddings = cached
                        self.b_data = None
                        log_print(self.log_box, f"⚠️ 旧缓存格式，请在勾选'强制重新向量化'后重新生成")
                        messagebox.showwarning("提示", "旧缓存格式，请勾选'强制重新向量化'重新生成")
                    self.vectorization_complete.set("blue")
                    self.vector_status_label.config(foreground=self.vectorization_complete.get())
                    self.update_match_btn_state()
                    log_print(self.log_box, "✅ 向量化完成")
                    return
                except Exception as e:
                    log_print(self.log_box, f"⚠️ 缓存加载失败，将重新生成：{str(e)}")
                    try:
                        os.remove(self.cache_file)
                    except:
                        pass

            log_print(self.log_box, f"正在生成向量（批量:{self.batch_size.get()}, 并发:{self.max_concurrency.get()}）...")

            all_texts = [d[4] for d in data]
            processed = [0]
            start_time = time.time()

            def update_progress(count):
                processed[0] += count
                pct = (processed[0] / len(all_texts)) * 100
                # 每25%打印一次进度
                if pct >= processed[0] * 100 // len(all_texts) // 25 * 25 or processed[0] == len(all_texts):
                    log_print(self.log_box, f"向量化进度: {pct:.0f}% ({processed[0]}/{len(all_texts)})")
                self.progress["value"] = (processed[0] / len(all_texts)) * 50
                self.root.update()

            api_key = self.api_key.get().strip()
            self.b_embeddings = get_embeddings_parallel(
                all_texts,
                api_key,
                batch_size=self.batch_size.get(),
                max_concurrency=self.max_concurrency.get(),
                progress_callback=update_progress,
                stop_check=lambda: self.stop_flag.is_set()
            )

            # 检查是否被停止
            if self.stop_flag.is_set():
                log_print(self.log_box, "❌ 向量化已停止")
                self.vectorizing.set(False)
                self.vectorize_btn.config(text="开始向量化")
                return

            if self.b_embeddings is not None and hasattr(self.b_embeddings, 'shape') and len(self.b_embeddings.shape) >= 2 and self.b_embeddings.shape[0] == len(data):
                # 确保缓存目录存在
                os.makedirs(VECTOR_CACHE_DIR, exist_ok=True)
                # 保存向量和B表数据到同一缓存文件
                cache_data = {'vectors': self.b_embeddings, 'data': data}
                np.save(self.cache_file, cache_data)
                log_print(self.log_box, f"✅ 向量已缓存: {self.b_embeddings.shape[0]} 条")
                self.progress["value"] = 100
                elapsed = time.time() - start_time
                log_print(self.log_box, f"✅ 存货向量生成并缓存完成（耗时:{elapsed:.1f}秒）")
                self.vectorization_complete.set("blue")
                self.vector_status_label.config(foreground=self.vectorization_complete.get())
                self.update_match_btn_state()
                log_print(self.log_box, "✅ 向量化完成")
                self.vectorizing.set(False)
                self.vectorize_btn.config(text="开始向量化")
                return
            else:
                log_print(self.log_box, "❌ 向量生成失败，未缓存")
                self.vectorizing.set(False)
                self.vectorize_btn.config(text="开始向量化")
        except Exception as e:
            log_print(self.log_box, f"❌ 加载失败：{str(e)}")
            self.vectorizing.set(False)
            self.vectorize_btn.config(text="开始向量化")

    def cos_sim(self, a, b):
        norm = np.linalg.norm(a) * np.linalg.norm(b, axis=1)
        return np.dot(a, b.T) / np.where(norm == 0, 1e-10, norm)

    def _get_cache_path(self, b_path=None):
        """获取向量缓存文件路径"""
        path = b_path or self.b_path.get()
        return os.path.join(VECTOR_CACHE_DIR, f"{os.path.basename(path)}.npy")

    def start_match(self):
        if not self.a_path.get():
            messagebox.showwarning("提示", "请选择要匹配的数据A表")
            return
        if self.b_embeddings is None or self.b_data is None:
            messagebox.showwarning("提示", "⚠️ 请先加载向量缓存后再执行匹配")
            return
        threading.Thread(target=self.run_match, daemon=True).start()

    def _load_b_data_from_excel(self):
        """从B表Excel加载数据（用于缓存加载后）"""
        path = self.b_path.get()
        if not path:
            return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            data = []

            # 自动查找列索引
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            col_to_idx = {}
            for i, cell in enumerate(header_row):
                if cell:
                    cell_str = str(cell).strip()
                    if "存货编码" in cell_str:
                        col_to_idx["code"] = i
                    elif "存货名称" in cell_str:
                        col_to_idx["name"] = i
                    elif "规格型号" in cell_str:
                        col_to_idx["spec"] = i
                    elif "尺寸" in cell_str:
                        col_to_idx["size"] = i

            if not all(k in col_to_idx for k in ["code", "name", "spec", "size"]):
                log_print(self.log_box, f"❌ B表缺少必要列")
                return

            for row in ws.iter_rows(min_row=2, values_only=True):
                code = str(row[col_to_idx["code"]]) if len(row) > col_to_idx["code"] else ""
                name = str(row[col_to_idx["name"]]) if len(row) > col_to_idx["name"] else ""
                spec = str(row[col_to_idx["spec"]]) if len(row) > col_to_idx["spec"] else ""
                size = str(row[col_to_idx["size"]]) if len(row) > col_to_idx["size"] else ""
                txt = clean_text(name + spec + size)
                if txt:
                    data.append((code, name, spec, size, txt))
            self.b_data = data
            log_print(self.log_box, f"✅ B表数据加载: {len(data)} 条")
        except Exception as e:
            log_print(self.log_box, f"❌ B表加载失败: {str(e)}")

    def run_match(self):
        log_print(self.log_box, "=" * 40)
        log_print(self.log_box, "开始AI匹配...")
        self.match_progress["value"] = 50
        api_key = self.api_key.get().strip()
        threshold = self.score_threshold.get()

        try:
            wb = openpyxl.load_workbook(self.a_path.get(), data_only=True)
            ws = wb.active
            a_list = []
            col_idx = ord(self.a_col.get().upper()) - 65
            for r in ws.iter_rows(min_row=2, values_only=True):
                txt = clean_text(str(r[col_idx]) if len(r) > col_idx else "")
                a_list.append(txt)
            log_print(self.log_box, f"✅ A表读取完成，共 {len(a_list)} 条")
        except:
            log_print(self.log_box, "❌ A表读取失败")
            return

        result = []
        pending_llm = []  # store LL.M tasks to run concurrently
        total = len(a_list)
        final_results = [None] * total
        self._pending_llm_details = {}
        match_count = 0
        fail_count = 0

        # Initialize final results container to ensure every item gets a definitive outcome
        final_results = [None] * total
        # Temporary storage for LL.M task details (for potential future diagnostics)
        self._pending_llm_details = {}

        for idx, a_txt in enumerate(a_list):
            if self.stop_flag.is_set():
                log_print(self.log_box, f"⚠️ 匹配中断，已处理 {idx}/{total} 条")
                break

            if not a_txt:
                result.append([a_txt, "", "", "", 0, "无内容"])
                continue

            emb = get_embeddings_with_retry([a_txt], api_key)
            if emb is None:
                log_print(self.log_box, f"❌ 第{idx+1}条：向量生成失败 - {a_txt[:30]}...")
                result.append([a_txt, "", "", "", 0, "向量失败"])
                fail_count += 1
                continue

            sim = self.cos_sim(emb[0], self.b_embeddings)
            top3_idx = np.argsort(sim)[::-1][:3]  # 取Top3
            raw_top3_data = [self.b_data[i] for i in top3_idx]
            # 去重：确保 B 候选中的名称/规格/尺寸组合不重复
            seen_keys = set()
            top3_data = []
            for item in raw_top3_data:
                key = (item[1], item[2], item[3])
                if key not in seen_keys:
                    seen_keys.add(key)
                    top3_data.append(item)
            top_sim = sim[top3_idx[0]]

            # Gate: Strategy B aggressive - skip LLM if there is at least 1 unique candidate
            top1_val = sim[top3_idx[0]] if len(top3_idx) > 0 else 0.0
            top2_val = sim[top3_idx[1]] if len(top3_idx) > 1 else 0.0
            gate_pass = False
            if (len(top3_data) >= 1) and (isinstance(top1_val, float) and math.isfinite(top1_val)):
                # 统一使用可配置门槛：仅当顶1分达到 sim_top1_th 时直接命中
                if top1_val >= self.sim_top1_th:
                    gate_pass = True

            # Log gate decision details for debugging
            log_print(self.log_box, f"第{idx+1}/{total}条: PASS={gate_pass}, top1={top1_val:.3f}, top2={top2_val:.3f}, diff={top1_val-top2_val:.3f}, candidates={len(top3_data)}")
            if gate_pass:
                first = top3_data[0]
                code = first[0]
                name = first[1]
                spec = first[2]
                size = first[3]
                ai_score = int(top1_val * 100)
                size_str = " " + size if size else ""
                # 新增：得分来源列，标记为 Cosine
                final_results[idx] = [a_txt, name, spec + size_str, code, ai_score, "Cosine", "已匹配"]
                log_print(self.log_box, f"Stage1: 直接匹配成功：{name} ({code}) 评分:{ai_score}")
                match_count += 1
            else:
                log_print(self.log_box, f"→ 第{idx+1}/{total}条：向量化相似度={top1_val:.3f}，Top2={top2_val:.3f}，进入LLM并发准备...")
                # Defer LL.M 调用，后续批量执行
                final_results[idx] = [a_txt, "", "", "", 0, "LLM", "LLM-pending"]
                cosine_score_for_this = int(top1_val * 100)
                pending_llm.append((idx, a_txt, top3_data, api_key, cosine_score_for_this))

            self.match_progress["value"] = 50 + ((idx + 1) / total) * 50

        # If there are pending LL.M tasks, run them concurrently using configured concurrency
        if pending_llm:
            from concurrent.futures import ThreadPoolExecutor, as_completed
            def llm_worker(item):
                idx, a_txt, top3_data, api_key, cosine_score_for_this = item
                try:
                    res = llm_find_best_from_top3(a_txt, top3_data, api_key, log_func=lambda msg: log_print(self.log_box, f"[LLM] {msg}"))
                    return (idx, res, cosine_score_for_this)
                except requests.exceptions.Timeout:
                    return (idx, ("timeout", 0))
                except Exception:
                    return (idx, (-1, None, None, None, None, 0))
            llm_results = []
            with ThreadPoolExecutor(max_workers=self.llm_concurrency.get()) as executor:
                futures = [executor.submit(llm_worker, item) for item in pending_llm]
                for f in as_completed(futures):
                    llm_results.append(f.result())
            for i, (idx, res, cosine_score) in enumerate(llm_results):
                # Update progress based on completion of LL.M tasks
                progress = 50 + ((i + 1) / max(1, len(llm_results))) * 50
                self.match_progress["value"] = progress
                self.root.update()
                if isinstance(res, tuple) and len(res) >= 2 and res[0] == "timeout":
                    final_results[idx] = [a_list[idx], "", "", "", 0, "LLM", "LLM超时"]
                elif isinstance(res, tuple) and len(res) >= 6:
                    best_idx, code, name, spec, size, llm_score = res
                    if best_idx >= 0:
                        size_str = " " + (size or "")
                        cosine_score_val = int(cosine_score)
                        llm_w = getattr(self, 'llm_weight', 0.5)
                        cos_w = getattr(self, 'cosine_weight', 0.5)
                        ai_score = int(round(llm_w * int(llm_score) + cos_w * cosine_score_val))
                        final_results[idx] = [a_list[idx], name, (spec + size_str).strip(), code, ai_score, "LLM", "已匹配"]
                    else:
                        final_results[idx] = [a_list[idx], "", "", "", 0, "LLM", "未达标"]
                else:
                    final_results[idx] = [a_list[idx], "", "", "", 0, "LLM", "未达标"]
        # 汇总输出：得分来源统计（Cosine/LLM）将在导出前计算并输出日志
        try:
            cosine_count = sum(1 for fr in final_results if isinstance(fr, list) and len(fr) >= 7 and fr[5] == "Cosine" and fr[6] == "已匹配")
            llm_count = sum(1 for fr in final_results if isinstance(fr, list) and len(fr) >= 7 and fr[5] == "LLM" and fr[6] == "已匹配")
            log_print(self.log_box, f"得分来源汇总: Cosine={cosine_count}, LLM={llm_count}")
        except Exception:
            pass
        out_path = os.path.join(os.path.dirname(self.a_path.get()), "AI匹配结果_最高相似.xlsx")
        # Ensure output directory exists to avoid permission errors on save
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        wb_out = openpyxl.Workbook()
        ws_out = wb_out.active
        ws_out.append(["A表内容", "B表名称", "B表规格", "B表编码", "AI得分", "得分来源", "状态"])
        # Write final_results in order; ensure 7 列并填充缺失值
        for i in range(total):
            fr = final_results[i]
            if fr is None:
                fr = [a_list[i], "", "", "", 0, "", "未达标"]
            if isinstance(fr, list) and len(fr) < 7:
                fr = fr + [""] * (7 - len(fr))
            ws_out.append(fr)
        try:
            wb_out.save(out_path)
        except Exception as e:
            log_print(self.log_box, f"❌ 写入主结果文件失败: {e}. 尝试输出备份文件。")
            backup_path = os.path.join(os.path.dirname(self.a_path.get()), f"AI匹配结果_最高相似_备份_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
            try:
                wb_out.save(backup_path)
                log_print(self.log_box, f"已输出备份文件：{backup_path}")
            except Exception as e2:
                log_print(self.log_box, f"备份输出失败：{e2}")

        self.match_progress["value"] = 100
        log_print(self.log_box, "=" * 40)
        log_print(self.log_box, f"✅ 匹配完成！共处理 {total} 条，匹配成功 {match_count} 条，失败 {fail_count} 条")
        log_print(self.log_box, f"✅ 结果已保存：{out_path}")
        messagebox.showinfo("完成", f"匹配完成！成功 {match_count} 条，失败 {fail_count} 条")

    def open_result_file(self):
        result_path = os.path.join(os.path.dirname(self.a_path.get()), "AI匹配结果_最高相似.xlsx")
        if os.path.exists(result_path):
            # 替换原来的 os.startfile(result_path)
            if sys.platform == "win32":
                os.startfile(result_path)
            else:
                subprocess.run(["open", result_path])
        else:
            messagebox.showwarning("提示", "结果文件不存在")

    def stop_match(self):
        self.stop_flag.set()  # 设置停止标志
        log_print(self.log_box, "❌ 匹配已停止")

    def create_widgets(self):
        # ==================== API 配置区 ====================
        frame_api = ttk.LabelFrame(self.root, text="API 配置")
        frame_api.place(x=20, y=10, width=860, height=70)

        ttk.Label(frame_api, text="API Key（阿里云 DashScope）：").place(x=10, y=15)
        self.api_key_entry = tk.Entry(frame_api, textvariable=self.api_key, width=45, show="*")
        self.api_key_entry.place(x=190, y=13)
        self.api_key_show = tk.BooleanVar(value=False)
        ttk.Checkbutton(frame_api, text="显示", variable=self.api_key_show, command=self.toggle_api_key_visibility).place(x=520, y=11)

        ttk.Button(frame_api, text="校验API", command=self.validate_api).place(x=580, y=11)
        ttk.Button(frame_api, text="保存配置", command=self.save_config).place(x=680, y=11)

        self.api_status = tk.StringVar(value="red")
        self.api_status_label = ttk.Label(frame_api, text="●", foreground=self.api_status.get(), font=("Arial", 14))
        self.api_status_label.place(x=800, y=10)

        # ==================== 第一步：B表处理区 ====================
        frame_step1 = ttk.LabelFrame(self.root, text="第一步：B表处理（底表向量化）")
        frame_step1.place(x=20, y=80, width=860, height=150)

        # B表文件选择
        ttk.Label(frame_step1, text="B表文件：").place(x=10, y=15)
        ttk.Entry(frame_step1, textvariable=self.b_path, width=40).place(x=80, y=13)
        ttk.Button(frame_step1, text="选择文件", command=self.select_b).place(x=360, y=11)

        # 向量化状态
        ttk.Label(frame_step1, text="向量化状态：").place(x=450, y=15)
        self.vector_status_label = ttk.Label(frame_step1, text="●", foreground=self.vectorization_complete.get(), font=("Arial", 14))
        self.vector_status_label.place(x=530, y=10)

        # 缓存选项
        ttk.Checkbutton(frame_step1, text="使用已有缓存", variable=self.use_cache, command=self.on_use_cache_toggle).place(x=10, y=50)
        ttk.Checkbutton(frame_step1, text="强制重新向量化", variable=self.force_revectorize).place(x=150, y=50)

        # 缓存文件选择（当使用已有缓存时显示）
        ttk.Label(frame_step1, text="缓存文件：").place(x=10, y=85)
        self.cache_combo = ttk.Combobox(frame_step1, width=12, state="readonly")
        self.cache_combo.place(x=80, y=83)
        self.cache_combo.bind('<<ComboboxSelected>>', self.on_cache_selected)

        # 批量大小和并发数
        ttk.Label(frame_step1, text="批量大小：").place(x=210, y=85)
        ttk.Combobox(frame_step1, textvariable=self.batch_size, values=[10, 20, 40], width=3, state="readonly").place(x=275, y=83)

        ttk.Label(frame_step1, text="并发数：").place(x=360, y=85)
        ttk.Combobox(frame_step1, textvariable=self.max_concurrency, values=[1, 2, 4, 8], width=3, state="readonly").place(x=420, y=83)

        # 向量化进度条
        ttk.Label(frame_step1, text="向量化进度：").place(x=480, y=85)
        self.progress = ttk.Progressbar(frame_step1, mode="determinate", length=200)
        self.progress.place(x=550, y=83, width=150, height=18)

        # 向量化按钮（开始/停止切换）
        self.vectorize_btn = ttk.Button(frame_step1, text="开始向量化", command=self.toggle_vectorization, width=15)
        self.vectorize_btn.place(x=750, y=78)




        # ==================== 第二步：A表处理区 ====================
        frame_step2 = ttk.LabelFrame(self.root, text="第二步：A表处理（执行匹配）")
        frame_step2.place(x=20, y=210, width=860, height=150)

        # A表文件选择
        ttk.Label(frame_step2, text="A表文件：").place(x=10, y=15)
        ttk.Entry(frame_step2, textvariable=self.a_path, width=40).place(x=80, y=13)
        ttk.Button(frame_step2, text="选择文件", command=self.select_a).place(x=360, y=11)
        


        # 阈值设置
        ttk.Label(frame_step2, text="相似度阈值：").place(x=450, y=15)
        ttk.Scale(frame_step2, variable=self.score_threshold, from_=0, to=100, orient=tk.HORIZONTAL, length=150).place(x=540, y=10)
        ttk.Label(frame_step2, textvariable=self.score_threshold).place(x=700, y=15)


        #LLM并发
        ttk.Label(frame_step2, text="LLM请求并发：", foreground="gray").place(x=20, y=58)
        if not hasattr(self, 'llm_concurrency_cb'):
            self.llm_concurrency_cb = ttk.Combobox(frame_step2, textvariable=self.llm_concurrency, values=[2, 4, 8, 10], width=6, state="readonly")
        self.llm_concurrency_cb.place(x=150, y=56)        

        # 匹配按钮
        self.match_btn = ttk.Button(frame_step2, text="开始匹配", command=self.start_match, width=15, state="disabled")
        self.match_btn.place(x=260, y=55)
        ttk.Button(frame_step2, text="停止运行", command=self.stop_match, width=15).place(x=410, y=55)

        # 提示
        ttk.Label(frame_step2, text="⚠️ 请先完成B表向量化后再执行匹配", foreground="gray").place(x=600, y=58)

        # 匹配进度条
        self.match_progress = ttk.Progressbar(frame_step2, mode="determinate")
        self.match_progress.place(x=60, y=93, width=790, height=20)
        ttk.Label(frame_step2, text="匹配进度：").place(x=10, y=95)

        
        frame_step3 = ttk.LabelFrame(self.root, text="第三步：结果查看")
        frame_step3.place(x=20, y=440, width=860, height=600)

        # 运行日志
        self.log_box = tk.Text(frame_step3, state=tk.DISABLED, font=("微软雅黑", 9))
        self.log_box.place(x=10, y=10, width=840, height=150)

        # 滚动条
        scrollbar = ttk.Scrollbar(frame_step3, command=self.log_box.yview)
        scrollbar.place(x=850, y=10, height=150)
        self.log_box.config(yscrollcommand=scrollbar.set)

        # 打开结果按钮
        ttk.Button(frame_step3, text="打开结果文件", command=self.open_result_file, width=15).place(x=10, y=180)

if __name__ == "__main__":
    root = tk.Tk()
    app = AISemanticMatchApp(root)
    root.mainloop()

